import time
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.shortcuts import render, redirect
from django.conf import settings
from django.http import HttpResponseRedirect, HttpResponse, Http404
from .models import EmployeeSheet, UploadMasterForProject, UploadMasterForEmployee, UploadMasterForClient

from .forms import EmployeeSheetForm, AddEmployeeForm, BulkUploadForm, BasicUploadForm, EmployeeSheetUploadForm,monthForm
from django.contrib import messages
from django.contrib.auth import authenticate
from django.contrib.auth import login
from django.contrib.auth import logout
from django.contrib.auth.decorators import user_passes_test, login_required
import os
from datetime import datetime
import openpyxl
from django.shortcuts import render_to_response
from openpyxl import load_workbook
from .choices import month,project
from django.core.files.storage import FileSystemStorage


@login_required(login_url='/login')
def home(request):
    data = UploadMasterForEmployee.objects.all()
    return render(request, 'home.html', {'data': data})


@login_required(login_url='/login')
def employee_data(request, e_data):
    employee = UploadMasterForEmployee.objects.get(pk=e_data)
    data = EmployeeSheet.objects.filter(employee_name_id=employee.id)
    return render(request, 'employee_data.html', {'data': data, 'employee': employee})


def add_employee(request):
    form = AddEmployeeForm()
    if request.method == "POST":
        form = AddEmployeeForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('/home')
    return render(request, 'add_employee.html', {'form': form})


@login_required(login_url='/login')
def add_record(request, e_id):
    employee = UploadMasterForEmployee.objects.get(pk=e_id)
    form = EmployeeSheetForm()
    if request.method == 'POST':
        form = EmployeeSheetForm(request.POST, request.FILES)
        if form.is_valid():
            form2 = form.save(commit=False)
            form2.employee_name = employee
            form2.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/employee_data/' + str(e_id))

    return render(request,'add_record.html',{'form':form})


# @login_required(login_url='/login')
# def upload_sheet(request, task_id):
#     form = EmployeeSheetUploadForm()
#     task_object = EmployeeSheet.objects.get(id=task_id)
#     e_id = task_object.employee_name_id
#     if request.POST:
#         form = EmployeeSheetUploadForm(request.POST, request.FILES, instance=task_object)
#         if form.is_valid():
#             form.save()
#         return HttpResponseRedirect('/employee_data/' + str(e_id))
#
#     return render(request, 'add_sheet.html', {'form': form})


@login_required(login_url='/login')
def download(request, e_id):
    try:
        task_object = EmployeeSheet.objects.get(id=e_id)
        path = task_object.sheet
        print(path)
        ee_id = task_object.employee_name_id
        file_path = os.path.join(settings.MEDIA_ROOT, str(path))
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        raise Http404
    except Exception as e:
        messages.error(request, "File Not Found")
        return HttpResponseRedirect('/employee_data/' + str(ee_id))


@login_required(login_url='/login')
def edit_task(request, k_id):
    data = EmployeeSheet.objects.get(pk=k_id)
    form = EmployeeSheetForm(instance=data)
    e_id = data.employee_name_id
    if request.POST:
        form = EmployeeSheetForm(request.POST, request.FILES, instance=data)
        if form.is_valid():
            form.save()
        else:
            print(form.errors)
            print("form is not valid")
        return HttpResponseRedirect('/employee_data/' + str(e_id))

    return render(request, 'edit_task.html', {'form': form})


def login_request(request):
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('UploadApp:upload')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid username or password.")
    form = AuthenticationForm()
    return render(request = request, template_name="login.html", context={"form":form})


def logout_request(request):
    logout(request)
    messages.info(request, "Logged out successfully!")
    return redirect("/login")

@login_required(login_url='/login')
def upload(request):
    return render(request, "upload.html")

@login_required(login_url='/login')
def basic_upload(request):
    if request.method == 'POST':
        form = BasicUploadForm(request.POST,request.FILES)

        if form.is_valid():
            sheet = form.cleaned_data['Month_calendar']

            print(sheet)
            files = request.FILES.get('basic_upload')
            name = files.name
            e_name = name.split('TimeSheet_')[1].split('.')[0]
            Name= e_name
            print("name",name)


            # path to media
            path_to_upload = os.path.join(settings.MEDIA_ROOT+'Myfile/'+e_name+'/'+sheet+'/')

            print('path_to_upload',path_to_upload)
            if os.path.exists(path_to_upload):
                print("Path exists")
            else:
                os.makedirs(path_to_upload)
            for obj, value in month:
                #print(obj, sheet)
                try:
                    if obj == sheet:
                        #print("Creating employee name in UploadMasterForEmployee", e_name)
                        data = UploadMasterForEmployee.objects.filter(employee_name=e_name)
                        e_id = ''
                        if data:
                            for i in data:
                                e_id = i.pk
                        else:
                            data = UploadMasterForEmployee.objects.create(employee_name=e_name)
                            e_id = data.pk

                        workbook = openpyxl.load_workbook(files)
                        worksheet = workbook[sheet]

                        ColNames = {}
                        Current = 0
                        for COL in worksheet.iter_cols(1, worksheet.max_column):
                            ColNames[COL[0].value] = Current
                            Current += 1
                        hours = []
                        minutes = []
                        time_list = []
                        project_hours = []
                        project_mins = []

                        for row_cells in worksheet.iter_rows(min_row=2):
                            #if row_cells[ColNames['Project']].value == selected_pro:
                            spent_time = row_cells[ColNames['Spent Time (Hrs)']].value
                            time_list.append(spent_time)

                            if spent_time:

                                if 'hrs' in spent_time:
                                    H = spent_time.replace('hrs', '')
                                    H = H.strip()
                                    if '.' in H:
                                        data = H.split('.')
                                        con_to_int = int(data[0])
                                        project_hours.append(con_to_int)
                                        if data[1] == '5':
                                            project_mins.append(30)
                                    elif ':' in H:
                                        data = H.split(':')
                                        con_to_int = int(data[0])
                                        project_hours.append(con_to_int)
                                        if data[1] == '5':
                                            project_mins.append(30)
                                    else:
                                        con_to_int = int(H)
                                        project_hours.append(con_to_int)

                                elif 'hr' in spent_time:
                                    H = spent_time.replace('hr', '')
                                    H = H.strip()
                                    if '.' in H:
                                        data = H.split('.')
                                        con_to_int = int(data[0])
                                        project_hours.append(con_to_int)
                                        if data[1] == '5':
                                            project_mins.append(30)
                                    elif ':' in H:
                                        data = H.split(':')
                                        con_to_int = int(data[0])
                                        project_hours.append(con_to_int)
                                        if data[1] == '5':
                                            project_mins.append(30)
                                    else:
                                        con_to_int = int(H)
                                        project_hours.append(con_to_int)

                                elif 'mins' in spent_time:
                                    M = spent_time.replace('mins', '')

                                    con = int(M)
                                    project_mins.append(con)


                            if row_cells[ColNames['Task Description ']].value:
                                upload_data = EmployeeSheet.objects.create(employee_name_id=e_id)
                                upload_em_data = EmployeeSheet.objects.get(pk=upload_data.pk)

                                date = row_cells[ColNames['Date']].value
                                upload_em_data.Month_calendar = sheet
                                #print('date1', date)
                                if date:
                                    req_date = datetime.strftime(date, '%Y-%m-%d')
                                    upload_em_data.Date = req_date

                                module = row_cells[ColNames['Module']].value
                                if module:
                                    upload_em_data.Module = module

                                task_d = row_cells[ColNames['Task Description ']].value
                                if task_d:
                                    upload_em_data.Task_description = task_d

                                spent_time = row_cells[ColNames['Spent Time (Hrs)']].value
                                if spent_time:
                                    upload_em_data.Spent_Time = spent_time
                                    if 'hrs' in spent_time:
                                        H = spent_time.replace('hrs', '')
                                        H = H.strip()
                                        if '.' in H:
                                            data = H.split('.')
                                            con_to_int = int(data[0])
                                            hours.append(con_to_int)
                                            if data[1] == '5':
                                                minutes.append(30)
                                        elif ':' in H:
                                            data = H.split(':')
                                            con_to_int = int(data[0])
                                            hours.append(con_to_int)
                                            if data[1] == '5':
                                                minutes.append(30)
                                        else:
                                            con_to_int = int(H)
                                            hours.append(con_to_int)
                                    elif 'hr' in spent_time:
                                        H = spent_time.replace('hr', '')
                                        H = H.strip()
                                        if '.' in H:
                                            data = H.split('.')
                                            con_to_int = int(data[0])
                                            hours.append(con_to_int)
                                            if data[1] == '5':
                                                minutes.append(30)
                                        elif ':' in H:
                                            data = H.split(':')
                                            con_to_int = int(data[0])
                                            hours.append(con_to_int)
                                            if data[1] == '5':
                                                minutes.append(30)
                                        else:
                                            con_to_int = int(H)
                                            hours.append(con_to_int)



                                        # print("hour list",hours)
                                    elif 'mins' in spent_time:
                                        M = spent_time.replace('mins', '')
                                        # M = M.strip()
                                        # print(M,"M")
                                        con = int(M)
                                        minutes.append(con)



                                remark = row_cells[ColNames['Remark']].value
                                if remark:
                                    upload_em_data.Remark = remark

                                client = row_cells[ColNames['Client Name ']].value
                                client_name = UploadMasterForClient.objects.filter(client_name=client)
                                get_client_name = ''
                                for i in client_name:
                                    get_client_name = i.pk

                                if get_client_name:
                                    upload_em_data.client_id = get_client_name

                                project = row_cells[ColNames['Project']].value


                                project_name = UploadMasterForProject.objects.filter(Project=project)
                                get_project_name = ''
                                for i in project_name:
                                    get_project_name = i.pk

                                if get_project_name:
                                    upload_em_data.project_id = get_project_name
                                    print(project_name,type(project_name))



                                upload_em_data.save()
                        sum_of_project_minutes = sum(project_mins)

                        sum_of_project_hrs = sum(project_hours)

                        project_sum = (sum_of_project_hrs) * 60 + sum_of_project_minutes

                        project_summation = project_sum / 60

                        print("project_summation", project_summation)
                        print(minutes, 'minutes')
                        sum_of_minutes = sum(minutes)

                        sum_of_hrs = sum(hours)

                        total_sum = (sum_of_hrs) * 60 + sum_of_minutes

                        summation = total_sum / 60
                        print(summation)

                        messages.info(request, "Uploaded successfully!")
                except:
                    messages.info(request, "Uploaded Failed")
                    return render(request, "basic_upload.html", {'form': form})

            return render(request, 'basic_upload.html', {'form': form})
    else:
        form = BasicUploadForm()
        return render(request, 'basic_upload.html', {'form': form})

@login_required(login_url='/login')
def bulk_upload(request):
    form = BulkUploadForm()
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            sheet = form.cleaned_data['month']
            #selected_pro = form.cleaned_data['project']
            files = request.FILES.getlist('bulk_upload')
            print('files', files,type(files))

            print(len(files))
            count= 0
            while count<=len(files):


                for f in files:
                    print('f', f)



                    for obj, value in month:

                       # print(obj)
                        if sheet in obj:
                            name = f.name
                            e_name = name.split('TimeSheet_')[1].split('.')[0]
                            print("e_name",e_name)

                            Name = e_name



                            # path to media
                            path_to_upload = os.path.join(settings.MEDIA_ROOT + 'Myfile/' + e_name + '/' + sheet + '/')

                            print('path_to_upload', path_to_upload)
                            if os.path.exists(path_to_upload):
                                print("Path exists")
                            else:
                                os.makedirs(path_to_upload)





                            data = UploadMasterForEmployee.objects.filter(employee_name=e_name)
                            e_id = ''
                            if data:
                                for i in data:
                                    e_id = i.pk
                            else:
                                data = UploadMasterForEmployee.objects.create(employee_name=e_name)
                                e_id = data.pk

                            workbook = openpyxl.load_workbook(f)
                            print(workbook,"workbook...........")

                            try:
                                worksheet = workbook[sheet]
                                ColNames = {}
                                Current = 0
                                for COL in worksheet.iter_cols(1, worksheet.max_column):
                                    ColNames[COL[0].value] = Current
                                    Current += 1

                                hours = []
                                minutes = []
                                time_list = []
                                project_hours = []
                                project_mins = []

                                for row_cells in worksheet.iter_rows(min_row=2):
                                    # if row_cells[ColNames['Project']].value == selected_pro:
                                    spent_time = row_cells[ColNames['Spent Time (Hrs)']].value
                                    time_list.append(spent_time)

                                    if spent_time:

                                        if 'hrs' in spent_time:
                                            H = spent_time.replace('hrs', '')
                                            H = H.strip()
                                            if '.' in H:
                                                data = H.split('.')
                                                con_to_int = int(data[0])
                                                project_hours.append(con_to_int)
                                                if data[1] == '5':
                                                    project_mins.append(30)
                                            elif ':' in H:
                                                data = H.split(':')
                                                con_to_int = int(data[0])
                                                project_hours.append(con_to_int)
                                                if data[1] == '5':
                                                    project_mins.append(30)
                                            else:
                                                con_to_int = int(H)
                                                project_hours.append(con_to_int)

                                        elif 'hr' in spent_time:
                                            H = spent_time.replace('hr', '')
                                            H = H.strip()
                                            if '.' in H:
                                                data = H.split('.')
                                                con_to_int = int(data[0])
                                                project_hours.append(con_to_int)
                                                if data[1] == '5':
                                                    project_mins.append(30)
                                            elif ':' in H:
                                                data = H.split(':')
                                                con_to_int = int(data[0])
                                                project_hours.append(con_to_int)
                                                if data[1] == '5':
                                                    project_mins.append(30)
                                            else:
                                                con_to_int = int(H)
                                                project_hours.append(con_to_int)







                                        elif 'mins' in spent_time:
                                            M = spent_time.replace('mins', '')

                                            con = int(M)
                                            project_mins.append(con)

                                    if row_cells[ColNames['Task Description ']].value:
                                        upload_data = EmployeeSheet.objects.create(employee_name_id=e_id)
                                        upload_em_data = EmployeeSheet.objects.get(pk=upload_data.pk)

                                        date = row_cells[ColNames['Date']].value
                                        upload_em_data.Month_calendar = sheet
                                        # print('date1', date)
                                        if date:
                                            req_date = datetime.strftime(date, '%Y-%m-%d')
                                            upload_em_data.Date = req_date

                                        module = row_cells[ColNames['Module']].value
                                        if module:
                                            upload_em_data.Module = module

                                        task_d = row_cells[ColNames['Task Description ']].value
                                        if task_d:
                                            upload_em_data.Task_description = task_d

                                        spent_time = row_cells[ColNames['Spent Time (Hrs)']].value
                                        if spent_time:
                                            upload_em_data.Spent_Time = spent_time
                                            if 'hrs' in spent_time:
                                                H = spent_time.replace('hrs', '')
                                                H = H.strip()
                                                if '.' in H:
                                                    data = H.split('.')
                                                    con_to_int = int(data[0])
                                                    hours.append(con_to_int)
                                                    if data[1] == '5':
                                                        minutes.append(30)
                                                elif ':' in H:
                                                    data = H.split(':')
                                                    con_to_int = int(data[0])
                                                    hours.append(con_to_int)
                                                    if data[1] == '5':
                                                        minutes.append(30)
                                                else:
                                                    con_to_int = int(H)
                                                    hours.append(con_to_int)

                                            elif 'hr' in spent_time:
                                                H = spent_time.replace('hr', '')
                                                H = H.strip()
                                                if '.' in H:
                                                    data = H.split('.')
                                                    con_to_int = int(data[0])
                                                    project_hours.append(con_to_int)
                                                    if data[1] == '5':
                                                        project_mins.append(30)
                                                elif ':' in H:
                                                    data = H.split(':')
                                                    con_to_int = int(data[0])
                                                    project_hours.append(con_to_int)
                                                    if data[1] == '5':
                                                        project_mins.append(30)
                                                else:
                                                    con_to_int = int(H)
                                                    project_hours.append(con_to_int)

                                                # print("hour list",hours)
                                            elif 'mins' in spent_time:
                                                M = spent_time.replace('mins', '')
                                                # M = M.strip()
                                                # print(M,"M")
                                                con = int(M)
                                                minutes.append(con)

                                        remark = row_cells[ColNames['Remark']].value
                                        if remark:
                                            upload_em_data.Remark = remark

                                        client = row_cells[ColNames['Client Name ']].value
                                        client_name = UploadMasterForClient.objects.filter(client_name=client)
                                        get_client_name = ''
                                        for i in client_name:
                                            get_client_name = i.pk

                                        if get_client_name:
                                            upload_em_data.client_id = get_client_name

                                        project = row_cells[ColNames['Project']].value

                                        project_name = UploadMasterForProject.objects.filter(Project=project)
                                        get_project_name = ''
                                        for i in project_name:
                                            get_project_name = i.pk

                                        if get_project_name:
                                            upload_em_data.project_id = get_project_name
                                            print(project_name, type(project_name))

                                        upload_em_data.save()
                                sum_of_project_minutes = sum(project_mins)

                                sum_of_project_hrs = sum(project_hours)

                                project_sum = (sum_of_project_hrs) * 60 + sum_of_project_minutes

                                project_summation = project_sum / 60

                                print("project_summation", project_summation)
                                print(minutes, 'minutes')
                                sum_of_minutes = sum(minutes)

                                sum_of_hrs = sum(hours)

                                total_sum = (sum_of_hrs) * 60 + sum_of_minutes

                                summation = total_sum / 60
                                print(summation)

                                messages.info(request, "Data Uploaded successfully for " + str(e_name))
                            except KeyError as exp:
                                messages.error(request, str(exp) + " for " + str(e_name))
                                return render(request, "bulk_upload.html", {'form': form})


                count+=1
                return render(request, "bulk_upload.html", {'form': form})

            return render(request, "report.html")


    else:
        return render(request, "bulk_upload.html", {'form': form})


def report(request):
    return render(request,'report.html')

def project_wise_report(request):
    if request.method == "POST":
        form = monthForm(request.POST)
        if form.is_valid():
            month = form.cleaned_data['month']



        project_list = project
        final_list = []
        for pro in project_list:
            pro_name = pro[0]
            hrs_list = []
            minutes_list = []
            up = UploadMasterForProject.objects.filter(Project = pro_name)
            for u in up:
                pro_id = u.pk
                print(u,"u")
                print(u.pk,"u.pk")
            result = EmployeeSheet.objects.filter(project_id = pro_id, Month_calendar = month)
            print(result,"result................................")
            for r in result:
                spent_time = r.Spent_Time
                if spent_time:
                    if 'hrs' in spent_time:
                        H = spent_time.replace('hrs', '')
                        H = H.strip()
                        if '.' in H:
                            data = H.split('.')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)
                        elif ':' in H:
                            data = H.split(':')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)



                    elif 'hr' in spent_time:
                        H = spent_time.replace('hr', '')
                        H = H.strip()
                        if '.' in H:
                            data = H.split('.')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)
                        elif ':' in H:
                            data = H.split(':')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)

                        else:
                            con_to_int = int(H)
                            hrs_list.append(con_to_int)


                    elif 'mins' in spent_time:
                        M = spent_time.replace('mins', '')

                        con = int(M)
                        minutes_list.append(con)

            sum_of_project_minutes = sum(minutes_list)

            sum_of_project_hrs = sum(hrs_list)

            project_sum = (sum_of_project_hrs) * 60 + sum_of_project_minutes

            project_summation = project_sum / 60
            print("project_summation",project_summation)

            final_list.append({pro_name:project_summation})

        return render(request,'project_wise_report.html',{'form':form,'final_list':final_list})

    else:
        form = monthForm()
        return render(request,'project_wise_report.html',{'form':form})

def user_wise_report(request):
    if request.method == "POST":
        form = monthForm(request.POST)
        if form.is_valid():
            month = form.cleaned_data['month']
            #selected_pro = form.cleaned_data['project']
        emp = UploadMasterForEmployee.objects.all()
        print("emp",emp)
        name_list = []
        emp_id = []
        for e in emp:
            name_list.append(e.employee_name)
            emp_id.append(e.pk)


        final_list = []
        for eid in emp_id:

            hrs_list = []
            minutes_list = []

            result = EmployeeSheet.objects.filter(employee_name=eid, Month_calendar = month)
            name = UploadMasterForEmployee.objects.get(id = eid)
            emp_name = name.employee_name
            for r in result:
                spent_time = r.Spent_Time
                # if row_cells[ColNames['Project']].value == selected_pro:
                if spent_time:
                    if 'hrs' in spent_time:
                        H = spent_time.replace('hrs', '')
                        H = H.strip()
                        if '.' in H:
                            data = H.split('.')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)
                        elif ':' in H:
                            data = H.split(':')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)
                        else:
                            con_to_int = int(H)
                            hrs_list.append(con_to_int)

                    elif 'hr' in spent_time:
                        H = spent_time.replace('hr', '')
                        H = H.strip()
                        if '.' in H:
                            data = H.split('.')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)
                        elif ':' in H:
                            data = H.split(':')
                            con_to_int = int(data[0])
                            hrs_list.append(con_to_int)
                            if data[1] == '5':
                                minutes_list.append(30)

                        else:
                            con_to_int = int(H)
                            hrs_list.append(con_to_int)


                    elif 'mins' in spent_time:
                        M = spent_time.replace('mins', '')

                        con = int(M)
                        minutes_list.append(con)

            sum_of_project_minutes = sum(minutes_list)

            sum_of_project_hrs = sum(hrs_list)

            project_sum = (sum_of_project_hrs) * 60 + sum_of_project_minutes

            project_summation = project_sum / 60

            final_list.append({emp_name:project_summation})

        return render(request,'user_wise_report.html',{'form':form,'final_list':final_list})



    else:
        form = monthForm()
        return render(request,'user_wise_report.html',{'form':form})

